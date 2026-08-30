"""Generate src/data/velvetTaxonomy.js from the authoritative workbook."""
from __future__ import annotations

import json
from collections import OrderedDict, defaultdict
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
XLSX = Path(r"c:\Users\jamal2000\Downloads\velvet_product_taxonomy.xlsx")
OUT = ROOT / "src" / "data" / "velvetTaxonomy.js"

OUTSIDE_IDS = {"815", "75", "76", "811", "810", "781"}


def split_bilingual(value: str) -> tuple[str, str]:
    text = str(value or "").strip()
    if " — " in text:
        ar, en = text.split(" — ", 1)
        return ar.strip(), en.strip()
    if " – " in text:
        ar, en = text.split(" – ", 1)
        return ar.strip(), en.strip()
    return text, text


def sheet_by_name(wb, needle: str):
    for name in wb.sheetnames:
        if needle in name:
            return wb[name]
    raise KeyError(needle)


def rows_from(ws):
    return list(ws.iter_rows(values_only=True))


def find_header(rows, required):
    for index, row in enumerate(rows):
        cells = [str(cell or "").strip() for cell in row]
        if all(item in cells for item in required):
            return index, cells
    raise ValueError(f"header not found for {required}")


def main() -> None:
    wb = load_workbook(XLSX, data_only=True, read_only=True)
    tree_rows = rows_from(sheet_by_name(wb, "شجرة"))
    import_rows = rows_from(sheet_by_name(wb, "الاستيراد"))

    tree_header_i, tree_header = find_header(
        tree_rows, ["Brand slug", "Main slug", "Leaf slug"]
    )
    import_header_i, import_header = find_header(
        import_rows, ["brand_slug", "product_id", "classification_status"]
    )

    brands: OrderedDict[str, dict] = OrderedDict()
    leaf_paths = set()
    main_slugs_by_brand: dict[str, set[str]] = defaultdict(set)
    leaf_slug_brands: dict[str, set[str]] = defaultdict(set)
    leaf_slug_paths: dict[str, set[str]] = defaultdict(set)

    for row in tree_rows[tree_header_i + 1 :]:
        data = {tree_header[i]: row[i] if i < len(row) else None for i in range(len(tree_header))}
        brand_slug = str(data.get("Brand slug") or "").strip()
        if not brand_slug:
            continue
        brand_code = str(data.get("كود البراند") or "").strip()
        brand_name = str(data.get("براند فيلفيت") or "").strip()
        main_label = str(data.get("الفئة الأساسية") or "").strip()
        leaf_label = str(data.get("الفئة الفرعية النهائية") or "").strip()
        main_slug = str(data.get("Main slug") or "").strip()
        leaf_slug = str(data.get("Leaf slug") or "").strip()
        count = data.get("عدد المنتجات")
        main_ar, main_en = split_bilingual(main_label)
        leaf_ar, leaf_en = split_bilingual(leaf_label)

        if brand_slug not in brands:
            brands[brand_slug] = {
                "code": brand_code,
                "slug": brand_slug,
                "name": {"en": brand_name, "ar": brand_name},
                "mainCategories": OrderedDict(),
            }
        brand = brands[brand_slug]
        mains = brand["mainCategories"]
        if main_slug not in mains:
            mains[main_slug] = {
                "code": "",
                "slug": main_slug,
                "nameAr": main_ar,
                "nameEn": main_en,
                "subcategories": OrderedDict(),
            }
        mains[main_slug]["subcategories"][leaf_slug] = {
            "code": "",
            "slug": leaf_slug,
            "nameAr": leaf_ar,
            "nameEn": leaf_en,
        }
        leaf_paths.add(f"{brand_slug}/{main_slug}/{leaf_slug}")
        main_slugs_by_brand[brand_slug].add(main_slug)
        leaf_slug_brands[leaf_slug].add(brand_slug)
        leaf_slug_paths[leaf_slug].add(f"{brand_slug}/{main_slug}/{leaf_slug}")

    product_map: OrderedDict[str, dict] = OrderedDict()
    status_counts = defaultdict(int)
    duplicate_ids = []
    outside = []

    for row in import_rows[import_header_i + 1 :]:
        data = {import_header[i]: row[i] if i < len(row) else None for i in range(len(import_header))}
        product_id = str(data.get("product_id") or "").strip()
        if not product_id:
            continue
        if product_id in product_map:
            duplicate_ids.append(product_id)
            continue
        brand_slug = str(data.get("brand_slug") or "").strip()
        main_slug = str(data.get("main_slug") or "").strip()
        leaf_slug = str(data.get("leaf_slug") or "").strip()
        status = str(data.get("classification_status") or "").strip()
        brand_code = str(data.get("brand_code") or "").strip()
        main_code = str(data.get("main_code") or "").strip()
        leaf_code = str(data.get("leaf_code") or "").strip()

        if brand_slug in brands:
            brands[brand_slug]["code"] = brands[brand_slug]["code"] or brand_code
            main = brands[brand_slug]["mainCategories"].get(main_slug)
            if main:
                main["code"] = main["code"] or main_code
                sub = main["subcategories"].get(leaf_slug)
                if sub:
                    sub["code"] = sub["code"] or leaf_code

        classified = bool(main_slug and leaf_slug)
        entry = {
            "brandSlug": brand_slug,
            "mainSlug": main_slug if classified else "",
            "subcategorySlug": leaf_slug if classified else "",
            "classificationStatus": status,
        }
        product_map[product_id] = entry
        status_counts[status] += 1
        if product_id in OUTSIDE_IDS or not classified:
            outside.append(product_id)

    brands_out = []
    for brand in brands.values():
        brands_out.append({
            "code": brand["code"],
            "slug": brand["slug"],
            "name": brand["name"],
            "mainCategories": [
                {
                    "code": main["code"],
                    "slug": main["slug"],
                    "nameAr": main["nameAr"],
                    "nameEn": main["nameEn"],
                    "subcategories": list(main["subcategories"].values()),
                }
                for main in brand["mainCategories"].values()
            ],
        })

    reused_leaf_slugs = sorted(
        slug for slug, paths in leaf_slug_paths.items() if len(paths) > 1
    )
    stats = {
        "brands": len(brands_out),
        "mainCategories": sum(len(brand["mainCategories"]) for brand in brands_out),
        "subcategories": len(leaf_paths),
        "productsMapped": len(product_map),
        "outsideTreeProducts": sorted(outside, key=lambda value: int(value) if value.isdigit() else value),
        "statusCounts": dict(status_counts),
        "duplicateProductIds": duplicate_ids,
        "reusedLeafSlugs": reused_leaf_slugs[:40],
        "reusedLeafSlugCount": len(reused_leaf_slugs),
    }

    js = (
        "// AUTO-GENERATED from velvet_product_taxonomy.xlsx. Do not edit by hand.\n"
        "// Regenerate with: python scripts/generate-velvet-taxonomy.py\n"
        "export const velvetTaxonomyStats = "
        + json.dumps(stats, ensure_ascii=False, indent=2)
        + ";\n\nexport const velvetTaxonomyBrands = "
        + json.dumps(brands_out, ensure_ascii=False, indent=2)
        + ";\n\nexport const productTaxonomyById = "
        + json.dumps(product_map, ensure_ascii=False, indent=2)
        + ";\n\nexport const outsideTaxonomyProductIds = "
        + json.dumps(stats["outsideTreeProducts"], ensure_ascii=False)
        + ";\n"
    )
    OUT.write_text(js, encoding="utf-8")
    Path("_tax_stats.json").write_text(json.dumps(stats, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(stats, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
